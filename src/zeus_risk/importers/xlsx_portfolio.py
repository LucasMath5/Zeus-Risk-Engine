"""Validated and resource-bounded XLSX adapter for portfolio positions."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import TypeAlias
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile

from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.cell.read_only import EmptyCell, ReadOnlyCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from zeus_risk.domain import ValidationIssue, ValidationSeverity
from zeus_risk.importers.models import ColumnMapping, ImportResult
from zeus_risk.importers.tabular import (
    TabularRecord,
    build_import_result,
    import_error,
    issue,
    map_columns,
)

_MAX_ARCHIVE_ENTRIES = 2_048
CellLike: TypeAlias = Cell | MergedCell | ReadOnlyCell | EmptyCell


@dataclass(frozen=True, slots=True)
class XlsxImportOptions:
    """Resource limits applied before and during synchronous workbook parsing."""

    max_file_size_bytes: int = 25 * 1024 * 1024
    max_uncompressed_size_bytes: int = 100 * 1024 * 1024
    max_rows: int = 25_000
    max_columns: int = 100

    def __post_init__(self) -> None:
        for field_name in (
            "max_file_size_bytes",
            "max_uncompressed_size_bytes",
            "max_rows",
            "max_columns",
        ):
            value = getattr(self, field_name)
            if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
                raise import_error(
                    "INVALID_XLSX_LIMIT",
                    "XLSX resource limits must be positive integers.",
                    field=field_name,
                )


class XlsxPortfolioImporter:
    """Import local XLSX workbooks without evaluating formulas or external links."""

    def __init__(self, options: XlsxImportOptions | None = None) -> None:
        self._options = options or XlsxImportOptions()

    @property
    def options(self) -> XlsxImportOptions:
        """Return immutable resource limits used by this importer."""

        return self._options

    def list_worksheets(self, path: str | Path) -> tuple[str, ...]:
        """Return workbook worksheet names in source order."""

        source_path = _validate_source(path, self.options)
        workbook = _load_workbook(source_path)
        try:
            return tuple(workbook.sheetnames)
        finally:
            workbook.close()

    def import_file(
        self,
        path: str | Path,
        *,
        worksheet_name: str | None = None,
        portfolio_name: str | None = None,
    ) -> ImportResult:
        """Import one explicitly selected or unambiguous worksheet."""

        source_path = _validate_source(path, self.options)
        source_name = str(source_path)
        workbook = _load_workbook(source_path)
        try:
            selected_name = _select_worksheet(
                tuple(workbook.sheetnames),
                worksheet_name,
                source_name,
            )
            worksheet = workbook[selected_name]
            resolved_portfolio_name = source_path.stem if portfolio_name is None else portfolio_name
            if not isinstance(resolved_portfolio_name, str) or not resolved_portfolio_name.strip():
                raise import_error(
                    "INVALID_PORTFOLIO_NAME",
                    "Portfolio name must be a non-empty string.",
                    field="name",
                    source_name=source_name,
                )
            return _import_worksheet(
                worksheet,
                source_name=source_name,
                portfolio_name=resolved_portfolio_name,
                options=self.options,
            )
        finally:
            workbook.close()


def _validate_source(path: str | Path, options: XlsxImportOptions) -> Path:
    if not isinstance(path, (str, Path)):
        raise import_error(
            "INVALID_FILE_PATH",
            "XLSX path must be a string or Path.",
            field="path",
        )

    source_path = Path(path)
    source_name = str(source_path)
    try:
        file_size = source_path.stat().st_size
    except FileNotFoundError as error:
        raise import_error(
            "FILE_NOT_FOUND",
            "Portfolio XLSX file was not found.",
            field="path",
            item=source_name,
            source_name=source_name,
        ) from error
    except OSError as error:
        raise import_error(
            "FILE_READ_ERROR",
            "Portfolio XLSX file could not be inspected.",
            field="path",
            item=source_name,
            source_name=source_name,
        ) from error

    if source_path.suffix.lower() != ".xlsx":
        raise import_error(
            "UNSUPPORTED_FILE_TYPE",
            "Only .xlsx workbooks are supported.",
            field="path",
            item=source_path.suffix or None,
            source_name=source_name,
        )
    if file_size > options.max_file_size_bytes:
        raise import_error(
            "XLSX_FILE_TOO_LARGE",
            "XLSX file exceeds the configured compressed-size limit.",
            field="path",
            item=str(file_size),
            source_name=source_name,
        )

    _validate_archive(source_path, options)
    return source_path


def _validate_archive(path: Path, options: XlsxImportOptions) -> None:
    source_name = str(path)
    try:
        with ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > _MAX_ARCHIVE_ENTRIES:
                raise import_error(
                    "XLSX_ARCHIVE_TOO_COMPLEX",
                    "XLSX archive contains too many entries.",
                    item=str(len(entries)),
                    source_name=source_name,
                )
            if any(entry.flag_bits & 0x1 for entry in entries):
                raise import_error(
                    "ENCRYPTED_XLSX_NOT_SUPPORTED",
                    "Encrypted XLSX workbooks are not supported.",
                    source_name=source_name,
                )

            uncompressed_size = sum(entry.file_size for entry in entries)
            if uncompressed_size > options.max_uncompressed_size_bytes:
                raise import_error(
                    "XLSX_ARCHIVE_TOO_LARGE",
                    "XLSX archive exceeds the configured uncompressed-size limit.",
                    item=str(uncompressed_size),
                    source_name=source_name,
                )
    except BadZipFile as error:
        raise import_error(
            "INVALID_XLSX",
            "File is not a valid XLSX archive.",
            source_name=source_name,
        ) from error
    except OSError as error:
        raise import_error(
            "FILE_READ_ERROR",
            "Portfolio XLSX file could not be read.",
            field="path",
            item=source_name,
            source_name=source_name,
        ) from error


def _load_workbook(path: Path) -> Workbook:
    source_name = str(path)
    try:
        return load_workbook(
            filename=path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except OSError as error:
        raise import_error(
            "FILE_READ_ERROR",
            "Portfolio XLSX file could not be read.",
            field="path",
            item=source_name,
            source_name=source_name,
        ) from error
    except (
        BadZipFile,
        InvalidFileException,
        DefusedXmlException,
        ParseError,
        KeyError,
        ValueError,
        EOFError,
    ) as error:
        raise import_error(
            "INVALID_XLSX",
            "Workbook structure is invalid or unsupported.",
            item=type(error).__name__,
            source_name=source_name,
        ) from error


def _select_worksheet(
    worksheet_names: tuple[str, ...],
    requested_name: str | None,
    source_name: str,
) -> str:
    if not worksheet_names:
        raise import_error(
            "EMPTY_WORKBOOK",
            "Workbook does not contain a worksheet.",
            source_name=source_name,
        )
    if requested_name is not None and (
        not isinstance(requested_name, str) or not requested_name.strip()
    ):
        raise import_error(
            "INVALID_WORKSHEET_NAME",
            "Worksheet name must be a non-empty string or None.",
            field="worksheet_name",
            source_name=source_name,
        )
    if requested_name is None:
        if len(worksheet_names) == 1:
            return worksheet_names[0]
        raise import_error(
            "WORKSHEET_SELECTION_REQUIRED",
            "Workbook contains multiple worksheets; choose one explicitly.",
            field="worksheet_name",
            item=", ".join(worksheet_names),
            source_name=source_name,
        )
    if requested_name not in worksheet_names:
        raise import_error(
            "WORKSHEET_NOT_FOUND",
            "Requested worksheet does not exist in the workbook.",
            field="worksheet_name",
            item=requested_name,
            source_name=source_name,
        )
    return requested_name


def _import_worksheet(
    worksheet: Worksheet,
    *,
    source_name: str,
    portfolio_name: str,
    options: XlsxImportOptions,
) -> ImportResult:
    max_row = worksheet.max_row or 0
    max_column = worksheet.max_column or 0
    if max_row > options.max_rows:
        raise import_error(
            "WORKSHEET_TOO_LONG",
            "Worksheet exceeds the configured physical-row limit.",
            field="worksheet",
            item=str(max_row),
            source_name=source_name,
        )
    if max_column > options.max_columns:
        raise import_error(
            "WORKSHEET_TOO_WIDE",
            "Worksheet exceeds the configured column limit.",
            field="worksheet",
            item=str(max_column),
            source_name=source_name,
        )

    row_iterator = enumerate(worksheet.iter_rows(), start=1)
    header: tuple[int, tuple[CellLike, ...]] | None = None
    for row_number, row in row_iterator:
        if not _cells_are_blank(row):
            header = (row_number, row)
            break
    if header is None:
        raise import_error(
            "EMPTY_WORKSHEET",
            "Selected worksheet does not contain a header.",
            field="worksheet",
            item=worksheet.title,
            source_name=source_name,
        )

    header_row_number, header_row = header
    header_cells = _trim_trailing_blank_cells(header_row)
    headers = tuple(
        _convert_header_cell(
            value=cell.value,
            data_type=cell.data_type,
            coordinate=f"{get_column_letter(column_index)}{header_row_number}",
            source_name=source_name,
        )
        for column_index, cell in enumerate(header_cells, start=1)
    )
    column_mappings, global_issues = map_columns(headers, source_name)

    records = tuple(
        _convert_data_row(
            row,
            row_number=row_number,
            column_mappings=column_mappings,
        )
        for row_number, row in row_iterator
        if not _cells_are_blank(row)
    )
    return build_import_result(
        source_name=source_name,
        portfolio_name=portfolio_name,
        column_mappings=column_mappings,
        records=records,
        global_issues=global_issues,
        worksheet_name=worksheet.title,
    )


def _convert_header_cell(
    *,
    value: object,
    data_type: str,
    coordinate: str,
    source_name: str,
) -> str:
    if data_type == "f":
        raise import_error(
            "FORMULA_IN_HEADER",
            "Worksheet headers cannot contain formulas.",
            field="header",
            item=coordinate,
            source_name=source_name,
        )
    if not isinstance(value, str):
        raise import_error(
            "INVALID_HEADER_CELL_TYPE",
            "Worksheet headers must contain text values.",
            field="header",
            item=coordinate,
            source_name=source_name,
        )
    return value


def _convert_data_row(
    cells: tuple[CellLike, ...],
    *,
    row_number: int,
    column_mappings: tuple[ColumnMapping, ...],
) -> TabularRecord:
    trimmed_cells = _trim_trailing_blank_cells(cells)
    values: list[str] = []
    raw_values: list[str] = []
    issues: list[ValidationIssue] = []

    for column_index, cell in enumerate(trimmed_cells, start=1):
        value = cell.value
        data_type = cell.data_type
        coordinate = f"{get_column_letter(column_index)}{row_number}"
        field = _field_for_column(column_index, column_mappings)
        parsed, raw, cell_issues = _convert_data_cell(
            value=value,
            data_type=data_type,
            coordinate=coordinate,
            field=field,
        )
        values.append(parsed)
        raw_values.append(raw)
        issues.extend(cell_issues)

    return TabularRecord(
        line_number=row_number,
        values=tuple(values),
        issues=tuple(issues),
        raw_values=tuple(raw_values),
    )


def _convert_data_cell(
    *,
    value: object,
    data_type: str,
    coordinate: str,
    field: str | None,
) -> tuple[str, str, tuple[ValidationIssue, ...]]:
    raw_value = _stringify_cell_value(value)
    if value is None:
        return "", "", ()
    if data_type == "f":
        return (
            "",
            raw_value,
            (
                issue(
                    ValidationSeverity.ERROR,
                    "FORMULA_NOT_ALLOWED",
                    "Formula cells are not accepted as portfolio inputs.",
                    field=field,
                    item=coordinate,
                ),
            ),
        )
    if data_type == "e":
        return (
            "",
            raw_value,
            (
                issue(
                    ValidationSeverity.ERROR,
                    "CELL_ERROR",
                    "Excel error cells are not accepted as portfolio inputs.",
                    field=field,
                    item=coordinate,
                ),
            ),
        )
    if isinstance(value, bool) or isinstance(value, (datetime, date, time)):
        return (
            "",
            raw_value,
            (
                issue(
                    ValidationSeverity.ERROR,
                    "UNSUPPORTED_CELL_TYPE",
                    "Boolean and date/time cells require explicit textual conversion.",
                    field=field,
                    item=coordinate,
                ),
            ),
        )
    if isinstance(value, (str, int, float, Decimal)):
        return raw_value, raw_value, ()
    return (
        "",
        raw_value,
        (
            issue(
                ValidationSeverity.ERROR,
                "UNSUPPORTED_CELL_TYPE",
                "Cell type is not supported for portfolio import.",
                field=field,
                item=coordinate,
            ),
        ),
    )


def _stringify_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _field_for_column(
    column_index: int,
    mappings: tuple[ColumnMapping, ...],
) -> str | None:
    if column_index > len(mappings):
        return None
    mapping = mappings[column_index - 1]
    return mapping.canonical_name or mapping.source_name


def _cells_are_blank(cells: tuple[CellLike, ...]) -> bool:
    return all(_cell_is_blank(cell) for cell in cells)


def _cell_is_blank(cell: CellLike) -> bool:
    value = cell.value
    return value is None or isinstance(value, str) and not value.strip()


def _trim_trailing_blank_cells(cells: tuple[CellLike, ...]) -> tuple[CellLike, ...]:
    last_value_index = len(cells)
    while last_value_index and _cell_is_blank(cells[last_value_index - 1]):
        last_value_index -= 1
    return cells[:last_value_index]
