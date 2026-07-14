"""Unit tests for the validated XLSX portfolio adapter."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import cast
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from zeus_risk.exceptions import PortfolioImportError
from zeus_risk.importers import ImportStatus, XlsxImportOptions, XlsxPortfolioImporter

HEADERS: list[object] = [
    "ticker",
    "quantity",
    "price",
    "asset_class",
    "currency",
    "sector",
]


def _save_workbook(path: Path, rows: list[list[object]], *, title: str = "Positions") -> Path:
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = title
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def test_imports_single_worksheet_with_numeric_and_text_cells(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "portfolio.xlsx",
        [
            HEADERS,
            ["PETR4", 10, 25.5, "equity", "BRL", "Energy"],
            ["VALE3", "5", "70.25", "equity", "BRL", "Materials"],
        ],
    )
    importer = XlsxPortfolioImporter()

    assert importer.list_worksheets(path) == ("Positions",)

    result = importer.import_file(path, portfolio_name="Brazil")

    assert result.source_name == str(path)
    assert result.worksheet_name == "Positions"
    assert result.encoding is None
    assert result.delimiter is None
    assert result.summary.total_rows == 2
    assert result.summary.valid_rows == 2
    assert result.summary.accepted_rows == 2
    assert result.portfolio is not None
    assert result.portfolio.name == "Brazil"
    assert result.positions[0].quantity == Decimal("10")
    assert result.positions[0].price == Decimal("25.5")
    assert result.rows[0].raw_fields[1].value == "10"


def test_uses_shared_aliases_and_preserves_physical_row(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "aliases.xlsx",
        [
            [],
            ["Ativo", "Quantidade", "Preço", "Classe de Ativo", "Moeda"],
            ["PETR4", -2, 30, "ação", "brl"],
        ],
    )

    result = XlsxPortfolioImporter().import_file(path)

    assert result.rows[0].line_number == 3
    assert result.rows[0].status is ImportStatus.WARNING
    assert result.rows[0].issues[0].code == "MISSING_OPTIONAL_SECTOR"
    assert result.positions[0].market_value == Decimal("-60")


def test_requires_explicit_selection_for_multiple_worksheets(tmp_path: Path) -> None:
    path = tmp_path / "multiple.xlsx"
    workbook = Workbook()
    first = cast(Worksheet, workbook.active)
    first.title = "First"
    first.append(HEADERS)
    first.append(["FIRST", 1, 10, "equity", "BRL", "Energy"])
    second = workbook.create_sheet("Second")
    second.append(HEADERS)
    second.append(["SECOND", 2, 20, "fund", "USD", "Fund"])
    workbook.save(path)
    workbook.close()
    importer = XlsxPortfolioImporter()

    assert importer.list_worksheets(path) == ("First", "Second")
    with pytest.raises(PortfolioImportError) as exc_info:
        importer.import_file(path)

    assert exc_info.value.issue.code == "WORKSHEET_SELECTION_REQUIRED"

    result = importer.import_file(path, worksheet_name="Second")

    assert result.worksheet_name == "Second"
    assert result.positions[0].instrument.ticker == "SECOND"


@pytest.mark.parametrize(
    ("worksheet_name", "expected_code"),
    [
        ("Missing", "WORKSHEET_NOT_FOUND"),
        ("", "INVALID_WORKSHEET_NAME"),
        (cast(str, 7), "INVALID_WORKSHEET_NAME"),
    ],
)
def test_rejects_invalid_worksheet_selection(
    tmp_path: Path,
    worksheet_name: str,
    expected_code: str,
) -> None:
    path = _save_workbook(
        tmp_path / "selection.xlsx",
        [HEADERS, ["PETR4", 1, 10, "equity", "BRL", "Energy"]],
    )

    with pytest.raises(PortfolioImportError) as exc_info:
        XlsxPortfolioImporter().import_file(path, worksheet_name=worksheet_name)

    assert exc_info.value.issue.code == expected_code


def test_rejects_formula_but_continues_with_other_rows(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "formula.xlsx",
        [
            HEADERS,
            ["FORMULA", 2, "=10+5", "equity", "BRL", "Energy"],
            ["VALID", 1, 20, "equity", "BRL", "Energy"],
        ],
    )

    result = XlsxPortfolioImporter().import_file(path)

    assert result.summary.error_rows == 1
    assert result.summary.accepted_rows == 1
    assert result.rows[0].issues[0].code == "FORMULA_NOT_ALLOWED"
    assert result.rows[0].issues[0].field == "price"
    assert result.rows[0].issues[0].item == "C2"
    assert result.rows[0].raw_fields[2].value == "=10+5"
    assert result.rows[1].status is ImportStatus.VALID
    assert result.is_partial


def test_rejects_boolean_date_and_excel_error_cells(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "types.xlsx",
        [
            HEADERS,
            ["BOOL", True, 10, "equity", "BRL", "Energy"],
            ["DATE", 1, datetime(2026, 7, 14), "equity", "BRL", "Energy"],
            ["ERROR", 1, 10, "equity", "#DIV/0!", "Energy"],
            ["VALID", 1, 10, "equity", "BRL", "Energy"],
        ],
    )

    result = XlsxPortfolioImporter().import_file(path)

    assert [row.status for row in result.rows] == [
        ImportStatus.ERROR,
        ImportStatus.ERROR,
        ImportStatus.ERROR,
        ImportStatus.VALID,
    ]
    assert result.rows[0].issues[0].code == "UNSUPPORTED_CELL_TYPE"
    assert result.rows[1].issues[0].code == "UNSUPPORTED_CELL_TYPE"
    assert result.rows[2].issues[0].code == "CELL_ERROR"
    assert result.summary.accepted_rows == 1


def test_applies_shared_duplicate_and_extra_field_rules(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "duplicates.xlsx",
        [
            HEADERS,
            ["PETR4", 1, 10, "equity", "BRL", "Energy"],
            ["petr4", 2, 20, "equity", "brl", "Energy"],
            ["VALE3", 1, 30, "equity", "BRL", "Materials", "unexpected"],
        ],
    )

    result = XlsxPortfolioImporter().import_file(path)

    assert result.rows[1].issues[-1].code == "DUPLICATE_POSITION"
    assert result.rows[2].issues[-1].code == "TOO_MANY_FIELDS"
    assert result.rows[2].raw_fields[-1].name == "__extra_1"
    assert result.summary.accepted_rows == 1


@pytest.mark.parametrize(
    ("headers", "expected_code"),
    [
        (["ticker", "quantity", "price"], "MISSING_REQUIRED_COLUMNS"),
        (
            ["ticker", "symbol", "quantity", "price", "asset_class", "currency"],
            "DUPLICATE_COLUMN_MAPPING",
        ),
        (["ticker", 2, "price", "asset_class", "currency"], "INVALID_HEADER_CELL_TYPE"),
        (["ticker", "=1+1", "price", "asset_class", "currency"], "FORMULA_IN_HEADER"),
    ],
)
def test_rejects_structural_header_problems(
    tmp_path: Path,
    headers: list[object],
    expected_code: str,
) -> None:
    path = _save_workbook(tmp_path / f"{expected_code}.xlsx", [headers, ["X", 1, 2]])

    with pytest.raises(PortfolioImportError) as exc_info:
        XlsxPortfolioImporter().import_file(path)

    assert exc_info.value.issue.code == expected_code


def test_rejects_empty_worksheet_and_header_without_rows(tmp_path: Path) -> None:
    empty_path = _save_workbook(tmp_path / "empty.xlsx", [])
    header_path = _save_workbook(tmp_path / "header.xlsx", [HEADERS])

    with pytest.raises(PortfolioImportError) as empty_error:
        XlsxPortfolioImporter().import_file(empty_path)
    with pytest.raises(PortfolioImportError) as no_rows_error:
        XlsxPortfolioImporter().import_file(header_path)

    assert empty_error.value.issue.code == "EMPTY_WORKSHEET"
    assert no_rows_error.value.issue.code == "NO_DATA_ROWS"


@pytest.mark.parametrize(
    ("filename", "content", "expected_code"),
    [
        ("portfolio.xls", b"legacy", "UNSUPPORTED_FILE_TYPE"),
        ("portfolio.xlsm", b"macro", "UNSUPPORTED_FILE_TYPE"),
        ("corrupt.xlsx", b"not a zip archive", "INVALID_XLSX"),
    ],
)
def test_rejects_unsupported_or_corrupt_files(
    tmp_path: Path,
    filename: str,
    content: bytes,
    expected_code: str,
) -> None:
    path = tmp_path / filename
    path.write_bytes(content)

    with pytest.raises(PortfolioImportError) as exc_info:
        XlsxPortfolioImporter().import_file(path)

    assert exc_info.value.issue.code == expected_code


def test_rejects_zip_without_valid_workbook_structure(tmp_path: Path) -> None:
    path = tmp_path / "invalid-structure.xlsx"
    with ZipFile(path, "w") as archive:
        archive.writestr("[Content_Types].xml", "<invalid")

    with pytest.raises(PortfolioImportError) as exc_info:
        XlsxPortfolioImporter().import_file(path)

    assert exc_info.value.issue.code == "INVALID_XLSX"


def test_reports_missing_file_and_directory_read_error(tmp_path: Path) -> None:
    missing = tmp_path / "missing.xlsx"
    directory = tmp_path / "folder.xlsx"
    directory.mkdir()

    with pytest.raises(PortfolioImportError) as missing_error:
        XlsxPortfolioImporter().import_file(missing)
    with pytest.raises(PortfolioImportError) as directory_error:
        XlsxPortfolioImporter().import_file(directory)

    assert missing_error.value.issue.code == "FILE_NOT_FOUND"
    assert directory_error.value.issue.code == "FILE_READ_ERROR"


def test_enforces_file_archive_and_worksheet_limits(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "limits.xlsx",
        [HEADERS, ["PETR4", 1, 10, "equity", "BRL", "Energy"]],
    )

    with pytest.raises(PortfolioImportError) as file_error:
        XlsxPortfolioImporter(XlsxImportOptions(max_file_size_bytes=1)).import_file(path)
    with pytest.raises(PortfolioImportError) as archive_error:
        XlsxPortfolioImporter(XlsxImportOptions(max_uncompressed_size_bytes=1)).import_file(path)
    with pytest.raises(PortfolioImportError) as row_error:
        XlsxPortfolioImporter(XlsxImportOptions(max_rows=1)).import_file(path)
    with pytest.raises(PortfolioImportError) as column_error:
        XlsxPortfolioImporter(XlsxImportOptions(max_columns=5)).import_file(path)

    assert file_error.value.issue.code == "XLSX_FILE_TOO_LARGE"
    assert archive_error.value.issue.code == "XLSX_ARCHIVE_TOO_LARGE"
    assert row_error.value.issue.code == "WORKSHEET_TOO_LONG"
    assert column_error.value.issue.code == "WORKSHEET_TOO_WIDE"


def test_rejects_invalid_limits_path_and_portfolio_name(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "boundary.xlsx",
        [HEADERS, ["PETR4", 1, 10, "equity", "BRL", "Energy"]],
    )

    with pytest.raises(PortfolioImportError) as limit_error:
        XlsxImportOptions(max_rows=cast(int, True))
    with pytest.raises(PortfolioImportError) as path_error:
        XlsxPortfolioImporter().import_file(cast(Path, 7))
    with pytest.raises(PortfolioImportError) as name_error:
        XlsxPortfolioImporter().import_file(path, portfolio_name="  ")

    assert limit_error.value.issue.code == "INVALID_XLSX_LIMIT"
    assert path_error.value.issue.code == "INVALID_FILE_PATH"
    assert name_error.value.issue.code == "INVALID_PORTFOLIO_NAME"
